"""Orchestration: generate data -> add noise -> train AE -> evaluate -> plot.

Revised per supervisor comments on AE_Mathematical_Foundations:
  - 7.4  Proper train/val/test split (60/20/20); val used for early
         stopping/model selection, test reserved for final reporting only.
  - 4.3  Optional per-trajectory normalization ablation (normalize.py).
  - 4.5  Configurable hidden width (model.py hidden_dim) for a leaner-AE
         comparison against the default 128.
  - 6.3  No-denoising baseline MSE (~= sigma^2) printed alongside test MSE.
  - Seeded runs so results are reproducible and comparable across seeds.

Single-config run:  python main.py
Batch/parallel run of many configs (see CONFIGS below): python main.py --batch
"""

import argparse
import os
from concurrent.futures import ProcessPoolExecutor, as_completed
from datetime import datetime
import numpy as np
import torch
import matplotlib
import matplotlib.pyplot as plt
from torchinfo import summary

from trajectory import generate_dataset
from noise import add_gaussian_noise
from model import TrajectoryAE
from train import train_model, evaluate_model
from normalize import compute_scale, normalize, denormalize
from visualize import (
    plot_clean_noisy_reconstructed, plot_training_curve,
    plot_per_axis_timeseries, plot_residuals, plot_latent_space_pca,
    plot_latent_space_pca_overlay, plot_trajectory_overlay,
    plot_mse_histogram, plot_snr_improvement, plot_noise_sweep,
)
from animate_preview import animate_sample, animate_train_test_comparison

# --- Default hyperparameters (used for `python main.py` single-config runs,
# and as the base that CONFIGS below overrides per-field for batch runs) ---
DEFAULTS = dict(
    N_TRAJECTORIES=300,   # total number of synthetic spiral trajectories to generate
    T=400,                # timesteps per trajectory
    NOISE_SIGMA=0.4,      # std dev of Gaussian noise added to each (x, y, z) point
    EMBEDDING_SIZE=4,     # size of the AE's compressed latent representation
    HIDDEN_DIMS=(128, 64, 32),  # encoder hidden-layer widths, tapered toward EMBEDDING_SIZE;
                                  # decoder mirrors this sequence in reverse
    EPOCHS=100,            # number of full passes over the training set
    BATCH_SIZE=32,         # trajectories per gradient update
    LR=1e-3,               # Adam learning rate (try 1e-4 to 1e-2 if loss stalls or diverges)
    TRAIN_SPLIT=0.6,       # fraction of trajectories used for training
    VAL_SPLIT=0.2,         # fraction used for validation (model selection / early stopping)
    # remaining 1 - TRAIN_SPLIT - VAL_SPLIT is held out as the final test set
    NORMALIZE=True,        # if True, normalize each trajectory by its own scale before
                            # the encoder and rescale the decoder output back afterwards
    SEED=50,                # random seed for numpy + torch, for reproducible / comparable runs
    SAMPLE_IDX=1,            # which TEST-split trajectory to plot/animate in reconstruction.png,
                              # per_axis_timeseries.png, residuals.png, trajectory_anim.mp4, etc.
                              # Valid range: 0 to (N_TRAJECTORIES - n_train - n_val - 1). If out of
                              # range, run_single_config() raises a clear error rather than
                              # silently clamping.
)

DATA_DIR = "data"       # folder where generated datasets are saved
RESULTS_ROOT = "results"  # parent folder for all per-run result subfolders

# Extra sigma levels to sweep for noise_sweep.png (Section 10 in the thesis).
# The config's own NOISE_SIGMA is trained/reported in full detail (loss curve,
# reconstruction, etc.); these additional levels are trained the same way (same
# architecture, epochs, split) purely to trace out MSE vs sigma. This is only
# used for single-config runs -- batch runs already sweep NOISE_SIGMA via
# CONFIGS, so PLOTS["noise_sweep"] is forced off there (see run_batch()).
NOISE_SWEEP_SIGMAS = [0.1, 0.2, 0.4, 0.6, 0.8, 1.0]
NOISE_SWEEP_SEEDS = [0, 1, 2]   # repeat each sigma over these seeds; report mean +/- std

# Which plots to generate. Core set covers: did training converge (loss_curve),
# one concrete before/after example (reconstruction), aggregate test-set error
# (mse_histogram), quantitative denoising gain (snr_improvement), and bottleneck
# representation quality / generalization (latent_space_pca_overlay). The rest
# are useful during development but redundant or non-essential for a final report.
PLOTS = {
    "trajectory_overlay": False,
    "loss_curve": True,
    "reconstruction": True,
    "reconstruction_worst": False,
    "per_axis_timeseries": False,
    "residuals": False,
    "mse_histogram": True,
    "snr_improvement": True,
    "latent_space_pca": False,
    "latent_space_pca_overlay": True,
    "noise_sweep": True,
}

# --- Batch grid: the set of configs run in parallel by `python main.py --batch`.
# Each dict overrides one or more DEFAULTS fields; everything else falls back
# to DEFAULTS. Currently: NOISE_SIGMA in {0.4, 0.6, 0.8, 1.0, 1.2} x
# EMBEDDING_SIZE in {4, 8, 16} x NORMALIZE in {True, False} = 30 configs.
CONFIGS = [
    {"NOISE_SIGMA": sigma, "EMBEDDING_SIZE": emb, "NORMALIZE": norm}
    for sigma in [0.4, 0.6, 0.8, 1.0, 1.2]
    for emb in [4, 8, 16]
    for norm in [True, False]
]

CONFIG_ANALYSIS_PATH = "configuration_ananlysis _MSE.xlsx"


def set_seed(seed):
    np.random.seed(seed)
    torch.manual_seed(seed)


def run_noise_sweep(cfg, clean, clean_train, clean_test, n_train, device):
    """For each sigma in NOISE_SWEEP_SIGMAS, train one fresh model per seed in
    NOISE_SWEEP_SEEDS (same architecture/epochs/split as cfg) and average the
    resulting test MSE across seeds.

    Returns:
        sigmas: NOISE_SWEEP_SIGMAS, sorted.
        mean_mses: mean test MSE per sigma, across seeds.
        std_mses: std dev of test MSE per sigma, across seeds.
    """
    T = cfg["T"]
    sigmas = sorted(NOISE_SWEEP_SIGMAS)
    mean_mses = []
    std_mses = []
    for sigma in sigmas:
        seed_mses = []
        for seed in NOISE_SWEEP_SEEDS:
            print(f"\n--- noise sweep: sigma = {sigma}, seed = {seed} ---")
            set_seed(seed)
            noisy_sweep = add_gaussian_noise(clean, sigma=sigma)
            noisy_sweep_flat = torch.tensor(noisy_sweep.reshape(clean.shape[0], -1), dtype=torch.float32)
            noisy_sweep_train = noisy_sweep_flat[:n_train]
            noisy_sweep_test = noisy_sweep_flat[n_train + int(clean.shape[0] * cfg["VAL_SPLIT"]):]

            model = TrajectoryAE(input_dim=T * 3, embedding_size=cfg["EMBEDDING_SIZE"],
                                  hidden_dims=cfg["HIDDEN_DIMS"])
            train_model(model, noisy_sweep_train, clean_train, epochs=cfg["EPOCHS"],
                        batch_size=cfg["BATCH_SIZE"], lr=cfg["LR"], device=device)
            sweep_test_loss, _ = evaluate_model(model, noisy_sweep_test, clean_test, device=device)
            print(f"sigma={sigma}  seed={seed}  test MSE={sweep_test_loss:.6f}")
            seed_mses.append(sweep_test_loss)
        mean_mses.append(float(np.mean(seed_mses)))
        std_mses.append(float(np.std(seed_mses)))
        print(f"sigma={sigma}  mean MSE={mean_mses[-1]:.6f}  std={std_mses[-1]:.6f}")
    return sigmas, mean_mses, std_mses


def append_config_analysis(cfg, timestamp, test_loss):
    """Append one row summarizing this run's config + result to the comparison
    spreadsheet (CONFIG_ANALYSIS_PATH), so results across runs stay comparable
    without digging through each results/ subfolder. Header is row 5,
    columns C-L; rows are appended starting at row 6.

    Not safe to call concurrently from multiple processes (would corrupt the
    file) -- in batch mode, only the parent process calls this, once per
    completed worker result.

    Existing rows plus the new one are re-sorted by (Noise_Sigma, Embedding
    size, Normalize) and rewritten with alternating group shading per sigma,
    so the sheet stays organized/readable after every run rather than just
    growing unsorted at the bottom.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    HEADER_ROW = 5
    FIRST_DATA_ROW = 6
    BAND_COLORS = ("FFF2CC", "D9E1F2")  # light yellow / light blue, alternating per sigma group

    wb = openpyxl.load_workbook(CONFIG_ANALYSIS_PATH)
    ws = wb["Sheet1"]

    rows = []
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(3, 13)]
        if any(v is not None for v in vals):
            rows.append(vals)

    rows.append([
        cfg["N_TRAJECTORIES"], timestamp, cfg["EMBEDDING_SIZE"], cfg["NOISE_SIGMA"],
        cfg["EPOCHS"], cfg["BATCH_SIZE"], round(test_loss, 6), str(cfg["HIDDEN_DIMS"]),
        cfg["LR"], cfg["NORMALIZE"],
    ])
    # Sort by sigma, then embedding size, then normalize (True before False).
    rows.sort(key=lambda v: (v[3], v[2], not v[9]))

    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        for c in range(3, 13):
            ws.cell(row=r, column=c).value = None
            ws.cell(row=r, column=c).fill = PatternFill(fill_type=None)

    current_row = FIRST_DATA_ROW
    prev_sigma = None
    color_idx = -1
    for vals in rows:
        sigma = vals[3]
        if sigma != prev_sigma:
            color_idx = (color_idx + 1) % len(BAND_COLORS)
            prev_sigma = sigma
            if current_row > FIRST_DATA_ROW:
                current_row += 1  # blank separator row between sigma groups
        fill = PatternFill(start_color=BAND_COLORS[color_idx], end_color=BAND_COLORS[color_idx], fill_type="solid")
        for c, v in enumerate(vals, start=3):
            cell = ws.cell(row=current_row, column=c, value=v)
            cell.fill = fill
        current_row += 1

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(3, 13):
        cell = ws.cell(row=HEADER_ROW, column=c)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")

    for c in range(3, 13):
        col_letter = get_column_letter(c)
        max_len = max(
            (len(str(ws.cell(row=r, column=c).value))
             for r in range(HEADER_ROW, current_row) if ws.cell(row=r, column=c).value is not None),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

    ws.freeze_panes = "C6"

    wb.save(CONFIG_ANALYSIS_PATH)
    print(f"Appended run to {CONFIG_ANALYSIS_PATH} (re-sorted, {len(rows)} total rows)")


def save_test_summary(cfg, run_dir, test_loss, baseline_mse, n_train, n_val, n_test, per_traj_mse):
    """Write a permanent, human-readable summary of this run's config and test
    results to test_summary.txt in run_dir.
    """
    best_idx = int(per_traj_mse.argmin())
    worst_idx = int(per_traj_mse.argmax())

    path = os.path.join(run_dir, "test_summary.txt")
    with open(path, "w") as f:
        f.write("Hyperparameters\n")
        f.write("---------------\n")
        f.write(f"N_TRAJECTORIES : {cfg['N_TRAJECTORIES']}\n")
        f.write(f"T              : {cfg['T']}\n")
        f.write(f"NOISE_SIGMA    : {cfg['NOISE_SIGMA']}\n")
        f.write(f"EMBEDDING_SIZE : {cfg['EMBEDDING_SIZE']}\n")
        f.write(f"HIDDEN_DIMS    : {cfg['HIDDEN_DIMS']}\n")
        f.write(f"EPOCHS         : {cfg['EPOCHS']}\n")
        f.write(f"BATCH_SIZE     : {cfg['BATCH_SIZE']}\n")
        f.write(f"LR             : {cfg['LR']}\n")
        f.write(f"SEED           : {cfg['SEED']}\n")
        f.write(f"NORMALIZE      : {cfg['NORMALIZE']}\n")
        f.write("\n")
        f.write("Split sizes\n")
        f.write("-----------\n")
        f.write(f"train : {n_train}\n")
        f.write(f"val   : {n_val}\n")
        f.write(f"test  : {n_test}\n")
        f.write("\n")
        f.write("Test results\n")
        f.write("------------\n")
        f.write(f"Test MSE               : {test_loss:.6f}\n")
        f.write(f"No-denoising baseline MSE: {baseline_mse:.6f}\n")
        f.write(f"Improvement over baseline: {baseline_mse / test_loss:.2f}x\n")
        f.write("\n")
        f.write("Per-trajectory stats\n")
        f.write("--------------------\n")
        f.write(f"Best  traj : idx={best_idx}, MSE={per_traj_mse[best_idx]:.6f}\n")
        f.write(f"Worst traj : idx={worst_idx}, MSE={per_traj_mse[worst_idx]:.6f}\n")
        f.write(f"Mean MSE   : {per_traj_mse.mean():.6f}\n")
        f.write(f"Std MSE    : {per_traj_mse.std():.6f}\n")

    print(f"Saved test summary to {path}")


def make_run_dir(cfg):
    """Create a fresh timestamped, labelled subfolder under RESULTS_ROOT for
    this run's plots, e.g. results/2026-07-22_18-30-05_N10_T650_sigma0.2_emb8/.
    """
    hidden_label = "-".join(str(d) for d in cfg["HIDDEN_DIMS"])
    label = (
        f"N{cfg['N_TRAJECTORIES']}_T{cfg['T']}_sigma{cfg['NOISE_SIGMA']}_emb{cfg['EMBEDDING_SIZE']}"
        f"_hid{hidden_label}_ep{cfg['EPOCHS']}_lr{cfg['LR']}_seed{cfg['SEED']}"
        f"{'_norm' if cfg['NORMALIZE'] else ''}"
    )
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    run_dir = os.path.join(RESULTS_ROOT, f"{timestamp}_{label}")
    os.makedirs(run_dir, exist_ok=True)
    return run_dir


def run_single_config(cfg, plots=None, open_files=True):
    """Run the full generate -> noise -> train -> evaluate -> plot pipeline
    for one hyperparameter configuration.

    Args:
        cfg: dict with all DEFAULTS keys (a full config, e.g. DEFAULTS itself
            or DEFAULTS overridden by one entry of CONFIGS).
        plots: which plots to generate (defaults to module-level PLOTS).
        open_files: if True, open generated plots/animations with the OS
            default viewer at the end (only sensible for a single foreground
            run -- batch/parallel runs should pass False).

    Returns:
        dict with run_dir, test_loss, and timestamp (run_dir's basename),
        for the caller to log/aggregate.
    """
    plots = PLOTS if plots is None else plots
    T = cfg["T"]
    N_TRAJECTORIES = cfg["N_TRAJECTORIES"]

    set_seed(cfg["SEED"])
    run_dir = make_run_dir(cfg)
    print(f"Saving results to {run_dir}/")

    def save(name):
        path = os.path.join(run_dir, name)
        plt.savefig(path, dpi=120)
        return path

    # 1. Generate ground-truth spiral trajectories and their noisy counterparts.
    #    clean/noisy shape: (N_TRAJECTORIES, T, 3)
    clean = generate_dataset(n_trajectories=N_TRAJECTORIES, T=T)
    noisy = add_gaussian_noise(clean, sigma=cfg["NOISE_SIGMA"])

    # 1b. Save the generated dataset to disk so it can be reused/inspected later.
    #     NOTE: in batch mode, concurrent workers overwrite this shared path --
    #     it's a "last writer wins" debug convenience, not per-run storage.
    #     Each run's own reconstructed_trajectories.npy is likewise shared;
    #     use run_dir's own saved arrays (per_traj_mse, test_summary.txt) as
    #     the authoritative per-run record instead.
    os.makedirs(DATA_DIR, exist_ok=True)
    np.save(os.path.join(DATA_DIR, "clean_trajectories.npy"), clean)
    np.save(os.path.join(DATA_DIR, "noisy_trajectories.npy"), noisy)
    print(f"Saved datasets to {DATA_DIR}/clean_trajectories.npy and {DATA_DIR}/noisy_trajectories.npy")

    # 1c. Overlay a sample of clean trajectories to show the dataset's spread.
    if plots["trajectory_overlay"]:
        plot_trajectory_overlay(clean, seed=0)
        save("trajectory_overlay.png")

    # 2. Flatten each (T, 3) trajectory into a single (T*3,) vector so it can
    #    be fed into the fully-connected autoencoder, and convert to tensors.
    clean_flat = torch.tensor(clean.reshape(N_TRAJECTORIES, -1), dtype=torch.float32)
    noisy_flat = torch.tensor(noisy.reshape(N_TRAJECTORIES, -1), dtype=torch.float32)

    # 3. Split into train/val/test sets (contiguous blocks: train, then val, then test).
    #    Val is used for early stopping / model selection; test is only touched once,
    #    at the end, for the final reported number.
    n_train = int(N_TRAJECTORIES * cfg["TRAIN_SPLIT"])
    n_val = int(N_TRAJECTORIES * cfg["VAL_SPLIT"])
    clean_train = clean_flat[:n_train]
    clean_val = clean_flat[n_train:n_train + n_val]
    clean_test = clean_flat[n_train + n_val:]
    noisy_train = noisy_flat[:n_train]
    noisy_val = noisy_flat[n_train:n_train + n_val]
    noisy_test = noisy_flat[n_train + n_val:]
    n_test = N_TRAJECTORIES - n_train - n_val

    # 3b. Optional per-trajectory normalization ablation (Section 4.3): scale each
    #     trajectory by its own noisy-input norm before the encoder ever sees it,
    #     and rescale the decoder's output back before computing any MSE/plots.
    #     Scale factors are computed independently per split, each from that
    #     split's own noisy data (never from clean data, which the model can't
    #     see at inference time).
    if cfg["NORMALIZE"]:
        scale_train = compute_scale(noisy_train, T)
        scale_val = compute_scale(noisy_val, T)
        scale_test = compute_scale(noisy_test, T)
        noisy_train_in, clean_train_in = normalize(noisy_train, scale_train), normalize(clean_train, scale_train)
        noisy_val_in, clean_val_in = normalize(noisy_val, scale_val), normalize(clean_val, scale_val)
        noisy_test_in, clean_test_in = normalize(noisy_test, scale_test), normalize(clean_test, scale_test)
    else:
        noisy_train_in, clean_train_in = noisy_train, clean_train
        noisy_val_in, clean_val_in = noisy_val, clean_val
        noisy_test_in, clean_test_in = noisy_test, clean_test

    # 4. Build the autoencoder: input/output dim = T*3 (flattened trajectory),
    #    bottleneck dim = EMBEDDING_SIZE, hidden widths tapered per HIDDEN_DIMS.
    device = "cuda" if torch.cuda.is_available() else "cpu"
    model = TrajectoryAE(input_dim=T * 3, embedding_size=cfg["EMBEDDING_SIZE"], hidden_dims=cfg["HIDDEN_DIMS"])
    n_params = sum(p.numel() for p in model.parameters() if p.requires_grad)

    print(f"Using device: {device}")
    print(f"Training samples  : {n_train}")
    print(f"Validation samples: {n_val}")
    print(f"Test samples      : {n_test}")
    print(f"Normalization     : {'per-trajectory (Section 4.3)' if cfg['NORMALIZE'] else 'off (raw coordinates)'}")
    print(model)
    print(f"\nTotal trainable parameters: {n_params:,}")

    # 4b. torchinfo layer-by-layer summary (input/output shapes, param counts).
    #     verbose=0 + manual print avoids UnicodeEncodeError on Windows consoles
    #     (cp1252) that can't render torchinfo's box-drawing characters.
    model_summary = summary(model, input_size=(cfg["BATCH_SIZE"], T * 3), device=device, verbose=0)
    print(str(model_summary).encode("ascii", "replace").decode("ascii"))

    print("\nStarting training...\n")

    # 5. Train the AE to map noisy trajectories -> clean trajectories (denoising),
    #    tracking validation loss every epoch alongside train loss. Validation
    #    (not test) drives early stopping / model selection.
    train_losses, val_losses = train_model(
        model, noisy_train_in, clean_train_in, epochs=cfg["EPOCHS"], batch_size=cfg["BATCH_SIZE"], lr=cfg["LR"],
        device=device, val_noisy=noisy_val_in, val_clean=clean_val_in)

    # 6. Evaluate final reconstruction quality on the held-out test set (touched
    #    only here, once, for the final reported number).
    test_loss_in, recon_in = evaluate_model(model, noisy_test_in, clean_test_in, device=device)

    # 6a. If normalization was used, rescale predictions/targets back to original
    #     units before reporting MSE or plotting, so numbers are comparable across
    #     NORMALIZE=True/False runs.
    if cfg["NORMALIZE"]:
        recon = denormalize(recon_in.cpu(), scale_test.cpu())
        test_loss = ((recon - clean_test) ** 2).mean().item()
    else:
        recon = recon_in.cpu()
        test_loss = test_loss_in

    # 6b. No-denoising baseline: the MSE an identity map (i.e. doing nothing)
    #     would achieve on the same test set. The relevant comparison for
    #     denoising performance is test_loss vs. this baseline, not vs. zero.
    baseline_mse = ((noisy_test - clean_test) ** 2).mean().item()
    print(f"Final test MSE          : {test_loss:.6f}")
    print(f"No-denoising baseline MSE (~sigma^2={cfg['NOISE_SIGMA']**2:.4f}): {baseline_mse:.6f}")
    print(f"Improvement over baseline: {baseline_mse / test_loss:.2f}x")

    # 6a. Save the denoised (reconstructed) test-set output to disk, shaped like
    #     the clean/noisy datasets: (n_test, T, 3). Always in original units.
    recon_array = recon.numpy().reshape(n_test, T, 3)
    np.save(os.path.join(DATA_DIR, "reconstructed_trajectories.npy"), recon_array)
    print(f"Saved denoised output to {DATA_DIR}/reconstructed_trajectories.npy")

    # 7. Plot the training/validation loss curve and save it.
    if plots["loss_curve"]:
        plot_training_curve(train_losses, val_losses)
        save("loss_curve.png")

    # 8. Pick one test trajectory, visualize clean vs noisy vs reconstructed, and save it.
    #    Also pick the worst-reconstructed test trajectory to inspect a bad case.
    #    All plotting below uses original-unit clean/noisy/recon tensors, regardless
    #    of whether NORMALIZE was used for training.
    per_traj_mse = ((recon.numpy() - clean_test.numpy()) ** 2).reshape(n_test, -1).mean(axis=1)
    worst_idx = int(per_traj_mse.argmax())

    save_test_summary(cfg, run_dir, test_loss, baseline_mse, n_train, n_val, n_test, per_traj_mse)

    sample_idx = cfg["SAMPLE_IDX"]
    if not (0 <= sample_idx < n_test):
        raise ValueError(f"SAMPLE_IDX={sample_idx} is out of range for the test split "
                          f"(0 to {n_test - 1}). Set SAMPLE_IDX to a valid test-set index.")
    clean_sample = clean_test[sample_idx].numpy().reshape(T, 3)
    noisy_sample = noisy_test[sample_idx].numpy().reshape(T, 3)
    recon_sample = recon[sample_idx].numpy().reshape(T, 3)

    if plots["reconstruction"]:
        plot_clean_noisy_reconstructed(clean_sample, noisy_sample, recon_sample)
        save("reconstruction.png")

    if plots["reconstruction_worst"]:
        clean_worst = clean_test[worst_idx].numpy().reshape(T, 3)
        noisy_worst = noisy_test[worst_idx].numpy().reshape(T, 3)
        recon_worst = recon[worst_idx].numpy().reshape(T, 3)
        plot_clean_noisy_reconstructed(clean_worst, noisy_worst, recon_worst)
        save("reconstruction_worst.png")

    # 8b. Per-axis time series and residual plots for the same sample.
    if plots["per_axis_timeseries"]:
        plot_per_axis_timeseries(clean_sample, noisy_sample, recon_sample)
        save("per_axis_timeseries.png")

    if plots["residuals"]:
        plot_residuals(clean_sample, recon_sample)
        save("residuals.png")

    # 8c. MSE distribution across the whole test set.
    if plots["mse_histogram"]:
        plot_mse_histogram(clean_test.numpy(), recon.numpy())
        save("mse_histogram.png")

    # 8d. SNR improvement bar chart for a random subset of test trajectories.
    #     worst_idx is forced into the sample and outlined, so the worst-case
    #     trajectory's SNR is always visible here, not left to chance.
    if plots["snr_improvement"]:
        plot_snr_improvement(clean_test.numpy(), noisy_test.numpy(), recon.numpy(), highlight_idx=worst_idx)
        save("snr_improvement.png")

    # 8e. Latent space visualization, colored by mean trajectory radius.
    #     Uses noisy_test_in (whatever the encoder actually consumes) so the
    #     embeddings match what the trained model produces.
    if plots["latent_space_pca"] or plots["latent_space_pca_overlay"]:
        with torch.no_grad():
            test_embeddings = model.encode(noisy_test_in.to(device)).cpu().numpy()

    if plots["latent_space_pca"]:
        clean_test_np = clean_test.numpy().reshape(n_test, T, 3)
        mean_radius = np.sqrt(clean_test_np[:, :, 0] ** 2 + clean_test_np[:, :, 1] ** 2).mean(axis=1)
        plot_latent_space_pca(test_embeddings, color_by=mean_radius, color_label='mean radius', split_name='test')
        save("latent_space_pca.png")

    # 8e-ii. Train vs test latent space overlay -- shows whether the bottleneck
    #        representation generalizes (test embeddings following the same
    #        structure as train) or diverges (a generalization gap).
    if plots["latent_space_pca_overlay"]:
        with torch.no_grad():
            train_embeddings = model.encode(noisy_train_in.to(device)).cpu().numpy()
        plot_latent_space_pca_overlay(train_embeddings, test_embeddings)
        save("latent_space_pca_overlay.png")

    # 8f. Noise-level sweep: reuse the same clean dataset/split/architecture,
    #     retrain at each sigma in NOISE_SWEEP_SIGMAS over NOISE_SWEEP_SEEDS
    #     seeds, and plot mean +/- std test MSE vs sigma. Off in batch mode --
    #     see run_batch().
    if plots["noise_sweep"]:
        sweep_sigmas, sweep_mean_mses, sweep_std_mses = run_noise_sweep(
            cfg, clean, clean_train, clean_test, n_train, device)
        plot_noise_sweep(sweep_sigmas, sweep_mean_mses, yerr=sweep_std_mses)
        save("noise_sweep.png")

    print(f"Saved plots to {run_dir}/ ({', '.join(name for name, on in plots.items() if on)})")

    # 9. Animate the same sample trajectory (clean vs noisy vs reconstructed) as an MP4,
    #    into the run folder. sample_idx indexes the full dataset for clean/noisy;
    #    recon_idx indexes reconstructed_trajectories.npy, which only covers the test split.
    #    Test starts after train + val (n_train + n_val), not directly after train.
    anim_path = os.path.join(run_dir, "trajectory_anim.mp4")
    animate_sample(data_dir=DATA_DIR, sample_idx=n_train + n_val + sample_idx, recon_idx=sample_idx,
                   out_path=anim_path)

    # 9b. Animate a "typical" train-split trajectory alongside a "typical"
    #     test-split trajectory (clean vs reconstructed for each), side by
    #     side. Trajectories are randomly generated (see trajectory.py), so
    #     picking by matching index number would pair two unrelated shapes;
    #     instead each sample is the one closest to its split's *median*
    #     per-trajectory MSE, giving a fair "typical case" comparison of
    #     reconstruction quality on seen (train) vs unseen (test) data.
    with torch.no_grad():
        _, train_recon_in = evaluate_model(model, noisy_train_in, clean_train_in, device=device)
    if cfg["NORMALIZE"]:
        train_recon = denormalize(train_recon_in.cpu(), scale_train.cpu())
    else:
        train_recon = train_recon_in.cpu()
    train_per_traj_mse = ((train_recon.numpy() - clean_train.numpy()) ** 2).reshape(n_train, -1).mean(axis=1)

    train_median_idx = int(np.argsort(train_per_traj_mse)[len(train_per_traj_mse) // 2])
    test_median_idx = int(np.argsort(per_traj_mse)[len(per_traj_mse) // 2])

    train_test_anim_path = os.path.join(run_dir, "train_test_comparison.mp4")
    animate_train_test_comparison(
        train_clean=clean_train[train_median_idx].numpy().reshape(T, 3),
        train_recon=train_recon[train_median_idx].numpy().reshape(T, 3),
        test_clean=clean_test[test_median_idx].numpy().reshape(T, 3),
        test_recon=recon[test_median_idx].numpy().reshape(T, 3),
        out_path=train_test_anim_path,
        train_mse=train_per_traj_mse[train_median_idx],
        test_mse=per_traj_mse[test_median_idx])

    if open_files:
        # 10. Open all generated plots/animation with the OS default viewer.
        plot_files = {
            "trajectory_overlay": "trajectory_overlay.png",
            "loss_curve": "loss_curve.png",
            "reconstruction": "reconstruction.png",
            "reconstruction_worst": "reconstruction_worst.png",
            "per_axis_timeseries": "per_axis_timeseries.png",
            "residuals": "residuals.png",
            "mse_histogram": "mse_histogram.png",
            "snr_improvement": "snr_improvement.png",
            "latent_space_pca": "latent_space_pca.png",
            "latent_space_pca_overlay": "latent_space_pca_overlay.png",
            "noise_sweep": "noise_sweep.png",
        }
        names = [name for key, name in plot_files.items() if plots[key]]
        names += ["trajectory_anim.mp4", "train_test_comparison.mp4", "test_summary.txt"]
        for name in names:
            os.startfile(os.path.join(run_dir, name))

    return {"run_dir": run_dir, "timestamp": os.path.basename(run_dir), "test_loss": test_loss}


def _run_worker(cfg):
    """Top-level, picklable entry point for ProcessPoolExecutor workers.
    Each worker process trains and evaluates one config in full isolation --
    its own dataset, its own model, its own run_dir -- and does NOT touch the
    shared CONFIG_ANALYSIS_PATH spreadsheet or open any files (that happens
    once, back in the parent process, in run_batch()).

    A non-interactive matplotlib backend is required here: worker processes
    have no display, and the default backend would error trying to open one.
    """
    matplotlib.use("Agg")
    plots = dict(PLOTS)
    plots["noise_sweep"] = False  # batch runs already sweep NOISE_SIGMA via CONFIGS
    result = run_single_config(cfg, plots=plots, open_files=False)
    return cfg, result


def run_batch(configs=None, max_workers=None):
    """Run every config in `configs` (defaults to CONFIGS), one process per
    config, up to max_workers at a time (defaults to os.cpu_count()).

    Each worker is fully independent (own dataset/model/run_dir); results are
    collected back here and appended to CONFIG_ANALYSIS_PATH one at a time as
    they complete, since concurrent writes to the same .xlsx would corrupt it.

    On a CUDA machine, multiple processes sharing one GPU usually contend for
    memory rather than genuinely speeding up -- if torch.cuda.is_available(),
    this falls back to running configs one at a time (max_workers=1) instead.
    """
    configs = CONFIGS if configs is None else configs
    full_configs = [{**DEFAULTS, **override} for override in configs]

    if torch.cuda.is_available() and max_workers is None:
        print("CUDA detected: running batch sequentially (max_workers=1) to avoid GPU contention.")
        max_workers = 1
    elif max_workers is None:
        max_workers = os.cpu_count() or 1

    print(f"Running {len(full_configs)} configs with max_workers={max_workers}...")

    completed, failed = 0, 0
    with ProcessPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(_run_worker, cfg): cfg for cfg in full_configs}
        for future in as_completed(futures):
            cfg = futures[future]
            try:
                cfg, result = future.result()
            except Exception as exc:
                failed += 1
                print(f"Config FAILED: sigma={cfg['NOISE_SIGMA']} emb={cfg['EMBEDDING_SIZE']} "
                      f"norm={cfg['NORMALIZE']} -- {exc}")
                continue
            append_config_analysis(cfg, result["timestamp"], result["test_loss"])
            completed += 1
            print(f"[{completed + failed}/{len(full_configs)}] Done: sigma={cfg['NOISE_SIGMA']} "
                  f"emb={cfg['EMBEDDING_SIZE']} norm={cfg['NORMALIZE']} -> test MSE={result['test_loss']:.6f}")

    print(f"\nBatch complete: {completed} succeeded, {failed} failed, out of {len(full_configs)}.")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--batch", action="store_true",
                         help="Run every config in CONFIGS in parallel instead of a single DEFAULTS run.")
    args = parser.parse_args()

    if args.batch:
        run_batch()
    else:
        run_single_config(DEFAULTS)


if __name__ == "__main__":
    main()
